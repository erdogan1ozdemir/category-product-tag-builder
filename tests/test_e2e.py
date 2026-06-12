"""Uçtan uca: raw veri → taxonomy → pools → tag → cross → export (ağsız, mock LLM)."""
import openpyxl
from core.brand_profile import BrandProfile
from core.pipeline import run_stage
from core.state import Workspace
from llm.bridge import MockBridge


def test_full_pipeline(tmp_path):
    brand = BrandProfile(name="Flormar", slug="flormar", sector="kozmetik")
    ws = Workspace("flormar", root=str(tmp_path)).ensure()
    ws.append_jsonl("products/products.jsonl",
                    {"id": "p1", "url": "https://x/1", "name": "Mat Ruj Kiremit",
                     "category": "Ruj", "description": "", "attributes": {"Bitiş": "Mat"}})
    ws.write_json("products/raw_facets.json",
                  {"Ruj": [{"Bitiş": ["Mat", "Parlak"], "Ton": ["Kiremit", "Nude"]}]})
    bridge = MockBridge({
        "taxonomy_proposal": {"groups": [{"group": "Bitiş"}, {"group": "Ton"}]},
        "pool_quality": {"remove": {}, "merge": {}},
        "product_tagging": {"products": [{"id": "p1", "tags": {"Ton": "Kiremit"}}]},
    })
    run_stage("taxonomy", brand, {}, root=str(tmp_path), bridge=bridge)
    out_pools = run_stage("pools", brand, {}, root=str(tmp_path), bridge=bridge)
    assert out_pools["built"] == ["Ruj"]
    pool = ws.read_json("pools/Ruj.json")
    pool["reviewed"] = True
    ws.write_json("pools/Ruj.json", pool)
    assert run_stage("tag", brand, {}, root=str(tmp_path), bridge=bridge) == 1
    out_cross = run_stage("cross", brand, {"seo": {"volume_threshold": 100}}, root=str(tmp_path))
    assert out_cross["combos"] >= 2
    assert out_cross["volume_source"] == "template"
    out = run_stage("export", brand, {}, root=str(tmp_path), targets=["excel"])
    wb = openpyxl.load_workbook(out["excel"])
    assert wb["Ürünler"].max_row == 2
    assert wb["Kombinasyonlar"].max_row >= 3
    row = ws.read_jsonl("tagged/tagged.jsonl")[0]
    # "Mat Ruj Kiremit" contains both "mat" and "kiremit" as literal substrings,
    # so deterministic matching covers Bitiş=Mat.  Ton=Kiremit is also matched
    # deterministically from the product name, so the LLM fill is never applied.
    assert row["tags"]["Bitiş"]["source"] == "deterministik"
