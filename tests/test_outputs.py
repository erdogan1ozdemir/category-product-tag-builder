import json
from unittest.mock import patch, MagicMock
import openpyxl
from core.state import Workspace
from facets.pool_builder import build_pool
from outputs.json_sink import export_json
from outputs.excel_sink import export_excel
from outputs.supabase_sink import SupabaseSink


def _ws(tmp_path):
    ws = Workspace("m", root=str(tmp_path)).ensure()
    ws.write_json("pools/Ruj.json", build_pool("Ruj", [{"Bitiş": ["Mat"]}], None))
    ws.append_jsonl("tagged/tagged.jsonl", {"id": "p1", "url": "u", "name": "Mat Ruj",
                                            "category": "Ruj",
                                            "tags": {"Bitiş": {"value": "Mat", "source": "deterministik"}}})
    ws.write_json("combos/combos.json",
                  [{"combo": "Mat Ruj", "volume": 900, "decision": "kategori"}])
    return ws


def test_export_json_consolidates(tmp_path):
    ws = _ws(tmp_path)
    path = export_json("m", ws)
    data = json.load(open(path, encoding="utf-8"))
    assert data["brand"] == "m"
    assert "Ruj" in data["pools"]
    assert data["tagged_count"] == 1


def test_export_excel_three_sheets_dynamic_columns(tmp_path):
    ws = _ws(tmp_path)
    path = export_excel("m", ws)
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Ürünler", "Havuz Özeti", "Kombinasyonlar"]
    urunler = wb["Ürünler"]
    headers = [c.value for c in urunler[1]]
    assert headers[:3] == ["Ürün Adı", "URL", "Kategori"]
    assert "Bitiş" in headers
    assert urunler.cell(row=2, column=headers.index("Bitiş") + 1).value == "Mat"


def test_supabase_sink_upserts(tmp_path):
    resp = MagicMock(status_code=201, text="")
    with patch("outputs.supabase_sink.requests.post", return_value=resp) as post:
        sink = SupabaseSink("https://x.supabase.co", "key")
        sink.upsert("pools", [{"brand_slug": "m", "category": "Ruj"}], conflict="brand_slug,category")
    url = post.call_args[0][0]
    assert url.endswith("/rest/v1/pools")
    assert post.call_args[1]["headers"]["apikey"] == "key"


def test_supabase_sink_chunks_large_batches(tmp_path):
    resp = MagicMock(status_code=201, text="")
    with patch("outputs.supabase_sink.requests.post", return_value=resp) as post:
        sink = SupabaseSink("https://x.supabase.co", "key")
        sink.upsert("product_tags", [{"brand_slug": "m", "product_id": str(i)} for i in range(1200)],
                    conflict="brand_slug,product_id")
    assert post.call_count == 3
    sizes = [len(c.kwargs["json"]) for c in post.call_args_list]
    assert sizes == [500, 500, 200]


def test_export_excel_sanitizes_illegal_chars(tmp_path):
    ws = Workspace("m", root=str(tmp_path)).ensure()
    ws.append_jsonl("tagged/tagged.jsonl", {"id": "p1", "url": "u", "name": "Kötü\x01Ad",
                                            "category": "Ruj", "tags": {}})
    path = export_excel("m", ws)
    wb = openpyxl.load_workbook(path)
    assert wb["Ürünler"].cell(row=2, column=1).value == "KötüAd"
