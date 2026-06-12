"""Genel Excel şablonu: Ürünler (dinamik facet kolonları), Havuz Özeti, Kombinasyonlar."""
import os
import re

import openpyxl
from openpyxl.styles import Font

from facets.pool_builder import pool_values

# openpyxl'in yasakladığı kontrol karakterleri (IllegalCharacterError önlemi)
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE as _ILLEGAL_RE
except ImportError:
    _ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def _clean(v):
    """String hücre değerlerinden yasadışı kontrol karakterlerini temizler."""
    return _ILLEGAL_RE.sub("", v) if isinstance(v, str) else v


def export_excel(brand_slug: str, ws) -> str:
    tagged = ws.read_jsonl("tagged/tagged.jsonl")
    combos = ws.read_json("combos/combos.json", default=[])
    pools = {}
    pools_dir = ws.path("pools")
    if os.path.isdir(pools_dir):
        for name in sorted(os.listdir(pools_dir)):
            if name.endswith(".json") and not name.startswith("_"):
                pools[name[:-5]] = ws.read_json(f"pools/{name}")

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    sh = wb.active
    sh.title = "Ürünler"
    facet_cols = sorted({g for r in tagged for g in r.get("tags", {})})
    headers = ["Ürün Adı", "URL", "Kategori"] + facet_cols
    sh.append(headers)
    for c in sh[1]:
        c.font = bold
    for r in tagged:
        row = [r.get("name"), r.get("url"), r.get("category")]
        row += [r.get("tags", {}).get(g, {}).get("value", "") for g in facet_cols]
        sh.append([_clean(x) for x in row])

    sh2 = wb.create_sheet("Havuz Özeti")
    sh2.append(["Kategori", "Grup", "Değer Sayısı", "Değerler"])
    for c in sh2[1]:
        c.font = bold
    for cat, pool in pools.items():
        for group, values in pool_values(pool).items():
            sh2.append([_clean(x) for x in [cat, group, len(values), ", ".join(values)]])

    sh3 = wb.create_sheet("Kombinasyonlar")
    sh3.append(["Kombinasyon", "Aranma Hacmi", "Karar"])
    for c in sh3[1]:
        c.font = bold
    for c in combos:
        sh3.append([_clean(x) for x in [c.get("combo"), c.get("volume"), c.get("decision")]])

    out = ws.path(f"exports/{brand_slug}.xlsx")
    wb.save(out)
    return out
